import openpyxl
import pandas as pd
import datetime as dt
from datetime import timedelta


def get_value(filename) -> dict:
    """
    Функция для получения точечно необходимой информации из Excel - файла
    """
    workbook = openpyxl.load_workbook(filename, data_only=True)

    sheet_parameters = workbook['Параметры']
    sheet_invest = workbook['Расчет инвест. затрат']
    sheet_chart_dl = workbook['Расчет графика ЛП']
    sheet_recoverable_cost = workbook['Возмещаемые затраты']

    data_dict = dict()

    # получаем проценты за финансирование
    total_percent = 0
    for row in sheet_invest.iter_rows(min_row=4, max_row=30):
        cell = row[3]
        if cell.value is not None:
            total_percent += cell.value
    data_dict['Итого проценты за финансирование'] = total_percent

    ## даты оплаты оплаты поставщику, размер оплаты поставщику, проценты за финансирование
    for row in sheet_invest.iter_rows(min_row=9, max_row=14):
        cell = row[1]
        if cell.value is not None and cell.value != 0:
            data_dict[sheet_invest.cell(row = cell.row, column=cell.column - 1).value + ' сумма'] = cell.value
            data_dict[sheet_invest.cell(row = cell.row, column=cell.column - 1).value + ' дата'] = sheet_invest.cell(row = cell.row, column=cell.column + 1).value.date()
            data_dict[sheet_invest.cell(row = cell.row, column=cell.column - 1).value + ' проценты'] = sheet_invest.cell(row = cell.row, column=cell.column + 2).value

    # получаем размер аванс
    data_dict['Размер аванса'] = sheet_invest['B6'].value

    # поучаем дату аванса
    data_dict['Дата аванса'] = sheet_invest['C6'].value.date()

    # получаем дату передачи
    data_dict['Дата передачи'] = sheet_invest['C3'].value.date()

    # получаем дату первого периода
    data_dict['Дата первого периода'] = sheet_chart_dl['B5'].value.date()

    # получаем проценты первого периода
    data_dict['Проценты первого периода'] = sheet_chart_dl['M5'].value

    # получаем сумму возмещаемых затрат
    total_cost = 0
    for row in sheet_recoverable_cost.iter_rows(min_row=3):
        cell = row[2]
        if cell.value is not None:
            total_cost += cell.value
    data_dict['Возмещаемые затраты'] = total_cost

    # получаем сумму ДЛ
    total_dl = 0
    for row in sheet_chart_dl.iter_rows(min_row=3):
        cell = row[13]
        if cell.value is not None:
            total_dl += cell.value
    data_dict['Сумма ДЛ'] = -total_dl

    # получаем сумму процентов за период лизинга
    total_percent = 0
    for row in sheet_chart_dl.iter_rows(min_row=3):
        cell = row[12]
        if cell.value is not None:
            total_percent += cell.value
    data_dict['Сумма процентов за период лизинга'] = total_percent

    # получаем сумму субсидии
    total_subs = 0
    for row in sheet_chart_dl.iter_rows(min_row=3):
        cell = row[14]
        if cell.value is not None:
            total_subs += cell.value
    data_dict['Сумма субсидии за лизинг'] = -total_subs

    # получаем данные с листа параметров: Итого_ДКП_С_НДС, Отсрочка ДЛ, Срок ДЛ в мес
    for row in sheet_parameters.iter_rows():
        cell = row[1]
        if cell.value == 'Отсрочка_ДЛ':
            data_dict['Отсрочка ДЛ'] = sheet_parameters.cell(row=cell.row, column=cell.column + 1).value
        elif cell.value == 'Срок_ДЛ_Мес':
            data_dict['Срок ДЛ'] = sheet_parameters.cell(row=cell.row, column=cell.column + 1).value
        elif cell.value == 'ИТОГО_ДКП_С_НДС':
            data_dict['ИТОГО_ДКП_С_НДС'] = sheet_parameters.cell(row=cell.row, column=cell.column + 1).value
    
    workbook.close()

    return data_dict
    
def answer(filename1, filename2=None):
    """
    Функция обработчик запроса из телеграм бота
    filename1: имя Excel - файла с значениями 
    filename2: имя второго Excel - файла для разбора инцидента с 2мя файлами
    """
    # TODO: сделать проверку на уникальность значений внутри словарей
    answer = 'Коллеги, добрый день,\n\nПо результатам рассмотрения инцидента могу сообщить следующее:\n\n'
    dict1 = get_value(filename1)

    # TODO: сделать для одного файла
    if filename2 is None:
        pass
    else:
        dict2 = get_value(filename2)
        answer += f'Разница в сумме ДЛ составила {round(dict2['Сумма ДЛ'] - dict1['Сумма ДЛ'])} руб. Из них:\n\n'

        # TODO: описать проценты за финансирование при нескольких оплатах поставщику
        if dict2['Итого проценты за финансирование'] != dict1['Итого проценты за финансирование']:
            for key in dict1:
                if key.endswith("дата"):
                    key_date = key
            date_diff = abs((dict1['Дата передачи'] - dict1[key_date]).days)
            answer += f'Проценты за финансирование\n\nВ предварительном расчете\n' \
                    f'Дата оплаты поставщику: {dict1.get(key_date)}\n'\
                    f'Дата передачи ПЛ: {dict1.get("Дата передачи")}\n' \
                    f'Разница между датой оплаты поставщику и датой передачи составляет: ' \
                    f'{date_diff} дней - начислено {dict1.get("Итого проценты за финансирование")} руб.\n\n'
            
            for key in dict2:
                if key.endswith("дата"):
                    key_date = key
            date_diff = abs((dict2['Дата передачи'] - dict2[key_date]).days)
            answer += f'В расчете по факту\n' \
                    f'Дата оплаты поставщику: {dict2.get(key_date)}\n'\
                    f'Дата передачи ПЛ: {dict2.get("Дата передачи")}\n' \
                    f'Разница между датой оплаты поставщику и датой передачи составляет: ' \
                    f'{date_diff} дней - начислено {dict2.get("Итого проценты за финансирование")} руб.\n\n'
        
        # TODO: описать изменение возмещаемых затрат
        if dict2['Возмещаемые затраты'] != dict1['Возмещаемые затраты']:
            answer += 'Возмещаемые затраты\n\n'

        # TODO: описать изменение аванса
        if dict2['Размер аванса'] != dict1['Размер аванса']:
            answer += 'Размер аванса\n\n'

        # TODO: описать изменение отсрочки ДЛ
        if dict2['Отсрочка ДЛ'] != dict1['Отсрочка ДЛ']:
            answer += f'Отсрочка ДЛ\n\nВ предварительном расчете отсрочка ДЛ составляет: {dict1["Отсрочка ДЛ"]} мес.\n'\
                    f'В расчете по факту: {dict2["Отсрочка ДЛ"]} мес.\n'
            if dict2['Отсрочка ДЛ'] > dict1['Отсрочка ДЛ']:
                answer += 'Отсрочка больше -> Инвест. затраты (основной долг клиента) погашается медленнее -> Начислено больше процентов за период лизинга'
            else:
                answer += 'Отсрочка меньше -> Инвест. затраты (основной долг клиента) погашается быстрее -> Начислено меньше процентов за период лизинга'
    
    return answer

filename_1 = 'по факту 302425785.xlsx'
filename_2 = 'предварительный 302425787.xlsx'
print(answer(filename1=filename_2, filename2=filename_1))

# first_class = ExcelFile(filename_1)
# dict1 = first_class.get_value()
# print(dict1)
# first_class.close_wb()

# sc_class = ExcelFile(filename_2)
# dict2 = sc_class.get_value()
# print(dict2)
# sc_class.close_wb()

# # dicts = [l, s]
# df1 = pd.DataFrame.from_dict(dict1, orient='index')
# df2 = pd.DataFrame.from_dict(dict2, orient='index')
# keys = []
# for i in dict1.keys():
#     keys.append(i)
# print(keys)
# df = pd.DataFrame(dict1, index=dict1.keys()).join(pd.DataFrame(dict2, index=dict1.keys()))
# df = pd.merge(df1, df2, on=keys, suffixes=['_df1', '_df2'])
# df = pd.concat([pd.DataFrame.from_dict(d, orient='index') for d in dicts], ignore_index=True)
# df = pd.concat([df1, df2], axis=1)
# # print(df)
# df.to_excel("data.xlsx", index='False')
